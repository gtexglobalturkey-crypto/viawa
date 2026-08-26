import json, os, sys, time, unicodedata, uuid
from urllib.parse import quote
from urllib.error import HTTPError
from urllib.request import Request, urlopen
import openpyxl

BASE = os.environ["SMOKE_SUPABASE_URL"].rstrip("/") + "/rest/v1"
KEY = os.environ["SMOKE_SERVICE_KEY"]
RUN = "SMOKE_" + uuid.uuid4().hex[:10].upper()
HEADERS = {"apikey": KEY, "Authorization": "Bearer " + KEY, "Content-Type": "application/json"}

def call(method, path, body=None, prefer=None):
    headers = dict(HEADERS)
    if prefer: headers["Prefer"] = prefer
    req = Request(BASE + path, data=None if body is None else json.dumps(body).encode(), headers=headers, method=method)
    try:
        with urlopen(req, timeout=60) as response:
            raw = response.read()
            return json.loads(raw) if raw else None
    except HTTPError as error:
        raise RuntimeError(error.read().decode("utf-8", "replace")) from error

def normalized(value):
    value = " ".join(str(value or "").strip().split()).replace("İ", "i").replace("I", "ı").lower()
    return unicodedata.normalize("NFC", value)

book = openpyxl.load_workbook(sys.argv[1], read_only=True, data_only=True)
sheet = book["VIAWA Import"]
headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
index = {name: position for position, name in enumerate(headers)}
source_rows = list(sheet.iter_rows(min_row=1171, max_row=1220, values_only=True))
created_company_ids, created_sector_ids, created_product_ids = [], [], []
users = call("GET", "/application_users?is_active=eq.true&select=id&limit=1")
if not users: raise RuntimeError("Staging has no active application user for contact ownership")
user_id = users[0]["id"]

def master_id(table, name, created_ids):
    key = normalized(name)
    existing = call("GET", f"/{table}?normalized_name=eq.{quote(key)}&select=id")
    if existing: return existing[0]["id"]
    row = call("POST", f"/{table}", {"name": str(name).strip(), "normalized_name": key}, "return=representation")
    created_ids.append(row[0]["id"])
    return row[0]["id"]

started = time.perf_counter()
try:
    for number, row in enumerate(source_rows, 1):
        company = call("POST", "/companies", {
            "company_name": f"{RUN} {row[index['Firma Ünvanı *']]}",
            "email": f"{RUN.lower()}.{number}@example.invalid",
            "phone": f"900000{number:05d}",
            "country": row[index["Ülke"]] or None,
            "city": row[index["Şehir"]] or None,
            "industry": row[index["Sektör 1"]] or None,
            "status": "lead",
        }, "return=representation")[0]
        company_id = company["id"]
        created_company_ids.append(company_id)
        sectors = [row[index[f"Sektör {slot}"]] for slot in range(1, 5) if row[index[f"Sektör {slot}"]]]
        products = [row[index[f"Ürün Grubu {slot}"]] for slot in range(1, 5) if row[index[f"Ürün Grubu {slot}"]]]
        for position, name in enumerate(dict.fromkeys(sectors), 1):
            call("POST", "/company_sectors", {"company_id": company_id, "sector_id": master_id("sectors", name, created_sector_ids), "position": position})
        for position, name in enumerate(dict.fromkeys(products), 1):
            call("POST", "/company_product_groups", {"company_id": company_id, "product_group_id": master_id("product_groups", name, created_product_ids), "position": position})
        for slot in range(1, 5):
            name = row[index[f"Kişi {slot} - Ad Soyad"]]
            if not name: continue
            parts = str(name).strip().split(maxsplit=1)
            call("POST", "/contacts", {"user_id": user_id, "company_id": company_id, "first_name": parts[0], "last_name": parts[1] if len(parts) > 1 else "", "title": row[index[f"Kişi {slot} - Ünvan"]] or None, "is_primary": False, "is_signatory": False})
    company_count = len(call("GET", f"/companies?company_name=like.{RUN}%25&select=id"))
    contact_count = len(call("GET", f"/contacts?company_id=in.({','.join(created_company_ids)})&select=id"))
    sector_relation_count = len(call("GET", f"/company_sectors?company_id=in.({','.join(created_company_ids)})&select=company_id"))
    product_relation_count = len(call("GET", f"/company_product_groups?company_id=in.({','.join(created_company_ids)})&select=company_id"))
    opportunity_count = len(call("GET", f"/opportunities?company_id=in.({','.join(created_company_ids)})&select=id"))
    page = call("POST", "/rpc/list_company_directory_page", {"p_page": 1, "p_page_size": 50, "p_search": RUN})
    result = {"run": RUN, "companies": company_count, "contacts": contact_count, "sector_relations": sector_relation_count, "product_relations": product_relation_count, "opportunities": opportunity_count, "directory_rows": len(page), "directory_total": int(page[0]["total_count"]) if page else 0, "seconds": round(time.perf_counter()-started, 3)}
    print(json.dumps(result, ensure_ascii=False))
    if company_count != 50 or opportunity_count != 0 or len(page) != 50 or result["directory_total"] != 50 or sector_relation_count == 0 or product_relation_count == 0:
        raise RuntimeError("Smoke verification failed")
finally:
    if created_company_ids:
        call("DELETE", f"/companies?id=in.({','.join(created_company_ids)})")
    for table, ids in (("sectors", created_sector_ids), ("product_groups", created_product_ids)):
        if ids: call("DELETE", f"/{table}?id=in.({','.join(ids)})")
    remaining = call("GET", f"/companies?company_name=like.{RUN}%25&select=id")
    print(json.dumps({"cleanup_remaining_companies": len(remaining)}))
